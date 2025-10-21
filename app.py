# app.py — AT&T → Limpieza (20 cols) + Estadísticos + Geocoding HTTP + Repositorio de Antenas (SQLite)
# v3.5 — Single-file listo para Streamlit Cloud
# - Reemplaza la caché en memoria por un REPOSITORIO PERSISTENTE (SQLite) de ubicaciones de antenas
# - Misma interfaz y Excel que la app de Limpieza
# - Sin dependencia de geopy

from __future__ import annotations
import io, os, re, sqlite3, tempfile, unicodedata
from datetime import datetime
from functools import lru_cache
from typing import Any, List, Dict, Tuple, Optional

import numpy as np
import pandas as pd
import requests
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

# --------- Opcional: PLUS CODE ----------
try:
    from openlocationcode import openlocationcode as olc
    HAS_OLC = True
except Exception:
    HAS_OLC = False

# ===========================
# CONFIG STREAMLIT
# ===========================
st.set_page_config(
    page_title="Go Mapper — AT&T → Limpieza (20 cols)",
    page_icon="📞",
    layout="wide",
)

st.title("📞 Go Mapper — Compilador AT&T (Limpieza + Repositorio de Antenas)")
st.write(
    "Convierte sábanas de AT&T al formato **Datos_Limpios (20 columnas)**, genera **PLUS_CODE** y **dirección**, "
    "calcula **estadísticos**, y mantiene un **repositorio propio** de ubicaciones de antenas en SQLite."
)

# ===========================
# PARÁMETROS UI
# ===========================
st.sidebar.header("Parámetros")
telefono_fijo = st.sidebar.text_input("Fijar columna 'Teléfono' (opcional)", value="", help="Si lo dejas vacío se usará NUM_A.")
remove_dups = st.sidebar.checkbox("Eliminar duplicados de DATOS por minuto (A/B), conservar la mayor duración", value=False)
show_preview = st.sidebar.checkbox("Mostrar preview", value=True)

st.sidebar.markdown("---")
st.sidebar.subheader("Repositorio de Antenas (persistente)")
DB_PATH = st.sidebar.text_input("Ruta del repositorio (SQLite)", value="antenas_repo.sqlite", help="Se crea automáticamente si no existe.")
PRECISION = st.sidebar.number_input("Precisión (decimales) para agrupar coords", min_value=0, max_value=7, value=6, step=1)
use_repo = st.sidebar.checkbox("Usar repositorio para resolver nombres antes de geocodificar", value=True)
update_repo = st.sidebar.checkbox("Actualizar repositorio con nuevas resoluciones", value=True)

st.sidebar.caption("Exporta un respaldo y/o importa tus curados.")
repo_export_btn = st.sidebar.button("⬇️ Exportar repositorio (CSV)")
repo_import_file = st.sidebar.file_uploader("⬆️ Importar/merge CSV al repositorio", type=["csv"]) 

# Geocoding HTTP: configuración básica
CONTACT_EMAIL = os.getenv("CONTACT_EMAIL", "contacto@example.com")
APP_VER       = "att-limpieza/3.5"
USER_AGENT    = f"{APP_VER} (+{CONTACT_EMAIL})"

NOMINATIM_URL = os.getenv("NOMINATIM_URL", "https://nominatim.openstreetmap.org").rstrip("/")
OPENCAGE_KEY   = os.getenv("OPENCAGE_API_KEY")   or st.secrets.get("OPENCAGE_API_KEY", "")
LOCATIONIQ_KEY = os.getenv("LOCATIONIQ_API_KEY") or st.secrets.get("LOCATIONIQ_API_KEY", "")

GEOCODER_TIMEOUT      = 20
MAX_UNIQUE_GEOCODES   = 800
COORD_PRECISION_CACHE = 6  # precisión para cache local de funciones

# ===========================
# PROGRESO
# ===========================
def progress_section(progress, pct, msg):
    progress.progress(min(max(int(pct), 0), 100), text=msg)

# ===========================
# UTILIDADES TXT/FECHAS/TIPO
# ===========================
def norm_text(x: Any) -> str:
    s = unicodedata.normalize("NFD", str(x or ""))
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    return re.sub(r"[^a-z0-9]+", " ", s).strip()

def parse_duration_to_seconds(val: Any) -> Optional[int]:
    if pd.isna(val): return None
    if isinstance(val, (int, float)) and not pd.isna(val): return int(round(float(val)))
    s = str(val).strip()
    if not s: return None
    if re.match(r"^\d{1,2}:[0-5]\d:[0-5]\d$", s):
        h, m, sec = s.split(":"); return int(h)*3600 + int(m)*60 + int(sec)
    if re.match(r"^[0-5]?\d:[0-5]\d$", s):
        m, sec = s.split(":"); return int(m)*60 + int(sec)
    s2 = re.sub(r"[^0-9]", "", s)
    return int(s2) if s2.isdigit() else None

VOZ_OUT = {"mo","moc","saliente","orig","out","originating","salida"}
VOZ_IN  = {"mt","mtc","entrante","term","in","terminating","entrada"}
MSG     = {"sms","mensaje","mensajes","2 vias","2vias","mms"}
DATA    = {"gprs","datos","data","internet","ps","pdp","packet"}
TRANSF  = {"transfer","desvio","desvío","call forward","cfu","cfb","cfnry","cfnr","cfnrc"}
VOICE_TOK = {"voz","llamada","call","moc","mtc"}

def derive_tipo(serv: Any, t_reg: Any, tipo_com: Any) -> Optional[str]:
    s = norm_text(serv); t = norm_text(t_reg); c = norm_text(tipo_com)
    if any(tok in s or tok in t or tok in c for tok in TRANSF): return "TRANSFER"
    if any(tok in s or tok in t or tok in c for tok in MSG):    return "MENSAJES 2 VÍAS"
    if any(tok in s or tok in t or tok in c for tok in DATA):   return "DATOS"
    is_voice = any(tok in s or tok in c for tok in VOICE_TOK) or ("call" in t) or ("moc" in t) or ("mtc" in t)
    if is_voice:
        if any(tok in t or tok in s or tok in c for tok in VOZ_OUT): return "VOZ SALIENTE"
        if any(tok in t or tok in s or tok in c for tok in VOZ_IN):  return "VOZ ENTRANTE"
        return "VOZ SALIENTE"
    if any(tok in t for tok in VOZ_OUT): return "VOZ SALIENTE"
    if any(tok in t for tok in VOZ_IN):  return "VOZ ENTRANTE"
    return None

def plus_code(lat, lon):
    if not HAS_OLC: return None
    try:
        latf, lonf = float(lat), float(lon)
        if not (-90 <= latf <= 90 and -180 <= lonf <= 180): return None
        return olc.encode(latf, lonf, codeLength=10)
    except Exception:
        return None

# ===========================
# DMS / COORDENADAS
# ===========================
def dms_to_decimal(value):
    if pd.isna(value):
        return None
    value = str(value).strip()
    if re.search(r"[°'\"NSEOWeo]", value, re.I):
        match = re.findall(r"(\d+)[°\s]+(\d+)?['\s]*([\d\.]+)?\"?\s*([NSEOWO])?", value, re.I)
        if not match:
            return None
        deg, minu, sec, hemi = match[0]
        deg = float(deg); minu = float(minu or 0); sec = float(sec or 0)
        dec = deg + minu/60 + sec/3600
        if hemi and hemi.upper() in ["S", "W", "O"]:
            dec = -dec
        return dec
    try:
        return float(value.replace(",", "."))
    except:
        return None

# ===========================
# HEADER SNIFF AT&T
# ===========================
REQ = {"NO","FECHA"}; ANY = {"DUR","DURACIÓN"}

def looks_like_header(vals: List[Any]) -> bool:
    up = {str(x).strip().upper() for x in vals if pd.notna(x)}
    return REQ.issubset(up) and len(ANY.intersection(up)) > 0

def read_any_with_sniff(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext in {".xlsx",".xlsm",".xls"}:
        raw = pd.read_excel(path, sheet_name=0, header=None, dtype=str)
        header_row = None
        for i in range(min(120, len(raw))):
            if looks_like_header(raw.iloc[i].tolist()):
                header_row = i; break
        if header_row is not None:
            headers = raw.iloc[header_row].tolist()
            df = raw.iloc[header_row+1:].copy()
            df.columns = headers
            return df.dropna(how="all", axis=1).dropna(how="all").reset_index(drop=True)
        return pd.read_excel(path, sheet_name=0)
    elif ext in {".csv",".txt"}:
        try:
            df = pd.read_csv(path, engine="python")
            if looks_like_header(list(df.columns)): return df
        except Exception: pass
        try:
            raw = pd.read_csv(path, engine="python", header=None, dtype=str)
            header_row = None
            for i in range(min(120, len(raw))):
                if looks_like_header(raw.iloc[i].tolist()):
                    header_row = i; break
            if header_row is not None:
                headers = raw.iloc[header_row].tolist()
                df = raw.iloc[header_row+1:].copy()
                df.columns = headers
                return df.dropna(how="all", axis=1).dropna(how="all").reset_index(drop=True)
        except Exception: pass
        return pd.read_csv(path, engine="python", encoding_errors="ignore")
    else:
        return pd.read_csv(path, engine="python", encoding_errors="ignore")

# ===========================
# AT&T → Datos_Limpios (20 columnas)
# ===========================
OUT_COLS = ['Teléfono','Tipo','Número A','Número B','Fecha','Hora','Duración (seg)','IMEI',
            'Latitud','Longitud','Azimuth','Latitud_raw','Longitud_raw','Azimuth_raw',
            'PLUS_CODE','PLUS_CODE_NOMBRE','Azimuth_deg','Datetime','Es_Duplicado','Cuenta_GrupoDup']

def transform_att_to_limpio(df: pd.DataFrame, telefono_fijo: Optional[str]) -> pd.DataFrame:
    out = pd.DataFrame(index=range(len(df)), columns=OUT_COLS)

    # Teléfono fijo -> si vacío, usar NUM_A
    out['Teléfono'] = (str(telefono_fijo).strip() if telefono_fijo else None) or df.get('NUM_A')

    # Tipo con SERV + T_REG + TIPO_COM
    serv = df.get('SERV'); treg = df.get('T_REG'); tipc = df.get('TIPO_COM')
    out['Tipo'] = [derive_tipo(s, t, c) for s,t,c in zip(
        serv if serv is not None else [None]*len(df),
        treg if treg is not None else [None]*len(df),
        tipc if tipc is not None else [None]*len(df)
    )] if (serv is not None or treg is not None or tipc is not None) else None

    # Número A / B (DEST o ID_DEST)
    out['Número A'] = df.get('NUM_A')
    out['Número B'] = df.get('DEST')
    if 'ID_DEST' in df.columns:
        mask_b = out['Número B'].isna() | (out['Número B'].astype(str).str.strip()=="")
        out.loc[mask_b,'Número B'] = df.loc[mask_b,'ID_DEST']

    # Fecha / Hora
    out['Fecha'] = df.get('FECHA')
    out['Hora']  = df.get('HORA')

    # Duración
    out['Duración (seg)'] = df.get('DUR').apply(parse_duration_to_seconds) if 'DUR' in df.columns else None

    # IMEI
    out['IMEI'] = df.get('NUM_A_IMEI')

    # Geo / Azimuth (raw y num)
    out['Latitud_raw'] = df.get('LATITUD')
    out['Longitud_raw'] = df.get('LONGITUD')
    out['Azimuth_raw'] = df.get('AZIMUTH')
    out['Latitud']  = pd.to_numeric(df.get('LATITUD'), errors='coerce') if 'LATITUD' in df.columns else None
    out['Longitud'] = pd.to_numeric(df.get('LONGITUD'), errors='coerce') if 'LONGITUD' in df.columns else None
    out['Azimuth'] = df.get('AZIMUTH')
    out['Azimuth_deg'] = pd.to_numeric(df.get('AZIMUTH'), errors='coerce') if 'AZIMUTH' in df.columns else None

    # PLUS_CODE y nombre (se rellena en geocoding)
    out['PLUS_CODE'] = [plus_code(lat, lon) if pd.notna(lat) and pd.notna(lon) else None
                        for lat, lon in zip(out['Latitud'], out['Longitud'])]
    out['PLUS_CODE_NOMBRE'] = None

    # Datetime (dayfirst)
    dt = pd.to_datetime(out['Fecha'].astype(str).str.strip() + " " + out['Hora'].astype(str).str.strip(),
                        errors="coerce", dayfirst=True)
    out['Datetime'] = dt

    # Flags duplicados iniciales
    out['Es_Duplicado'] = False
    out['Cuenta_GrupoDup'] = 1
    return out[OUT_COLS]

# ===========================
# DEDUPE DATOS (minuto)
# ===========================
def dedupe_datos_by_minute(df: pd.DataFrame):
    if df is None or df.empty or "Tipo" not in df.columns:
        return df, pd.DataFrame(), 0
    if not {"Número A","Número B"}.issubset(df.columns):
        return df, pd.DataFrame(), 0
    mask = df["Tipo"].astype(str).str.upper().eq("DATOS")
    otros = df.loc[~mask].copy()
    datos = df.loc[mask].copy()
    if "Datetime" in datos.columns and pd.api.types.is_datetime64_any_dtype(datos["Datetime"]):
        datos["__t__"] = datos["Datetime"].dt.floor("min")
    else:
        datos["__t__"] = datos["Hora"].astype(str).str[:5] if "Hora" in datos.columns else ""
    subset = ["Número A","Número B","__t__"]
    if "Duración (seg)" in datos.columns:
        datos = datos.sort_values("Duración (seg)", ascending=False)
    dup_extras = datos.duplicated(subset=subset, keep="first")
    duplicados_df = datos[dup_extras].copy()
    datos = datos[~dup_extras].drop(columns=["__t__"], errors="ignore")
    out = pd.concat([otros, datos], ignore_index=True)
    return out, duplicados_df, len(duplicados_df)

# ===========================
# REPOSITORIO DE ANTENAS (SQLite)
# ===========================
SCHEMA = """
CREATE TABLE IF NOT EXISTS antenas_repo (
  lat REAL NOT NULL,
  lon REAL NOT NULL,
  lat_round REAL NOT NULL,
  lon_round REAL NOT NULL,
  precision INTEGER NOT NULL,
  plus_code TEXT,
  nombre TEXT,
  fuente TEXT,
  confianza INTEGER,
  primera_vez TEXT,
  ultima_vez TEXT,
  veces INTEGER DEFAULT 1,
  PRIMARY KEY (lat_round, lon_round, precision)
);
"""

def repo_connect(path: str):
    con = sqlite3.connect(path)
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute(SCHEMA)
    con.commit()
    return con

def repo_fetch_map(con: sqlite3.Connection, precision: int) -> Dict[Tuple[float,float], Dict[str, Any]]:
    rows = con.execute(
        "SELECT lat_round, lon_round, nombre, plus_code, fuente, confianza, veces FROM antenas_repo WHERE precision=?",
        (precision,)
    ).fetchall()
    m: Dict[Tuple[float,float], Dict[str, Any]] = {}
    for r in rows:
        m[(float(r[0]), float(r[1]))] = {
            "nombre": r[2] or "",
            "plus_code": r[3] or "",
            "fuente": r[4] or "",
            "confianza": r[5] if r[5] is not None else None,
            "veces": r[6] if r[6] is not None else 0,
        }
    return m

def repo_upsert_many(con: sqlite3.Connection, items: List[Dict[str, Any]]):
    now = datetime.utcnow().isoformat(timespec="seconds")
    for it in items:
        con.execute(
            """
            INSERT INTO antenas_repo(lat,lon,lat_round,lon_round,precision,plus_code,nombre,fuente,confianza,primera_vez,ultima_vez,veces)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,1)
            ON CONFLICT(lat_round,lon_round,precision) DO UPDATE SET
              nombre=COALESCE(excluded.nombre, antenas_repo.nombre),
              plus_code=COALESCE(excluded.plus_code, antenas_repo.plus_code),
              fuente=COALESCE(excluded.fuente, antenas_repo.fuente),
              confianza=COALESCE(excluded.confianza, antenas_repo.confianza),
              ultima_vez=?,
              veces=antenas_repo.veces+1
            """,
            (
                it.get("lat"), it.get("lon"), it.get("lat_round"), it.get("lon_round"), it.get("precision"),
                it.get("plus_code"), it.get("nombre"), it.get("fuente"), it.get("confianza"),
                now, now, now
            )
        )
    con.commit()

def repo_export_df(con: sqlite3.Connection) -> pd.DataFrame:
    return pd.read_sql_query("SELECT * FROM antenas_repo ORDER BY precision DESC, lat_round, lon_round", con)

def repo_import_df(con: sqlite3.Connection, df: pd.DataFrame):
    req_cols = {"lat","lon","lat_round","lon_round","precision","nombre"}
    if not req_cols.issubset(set(df.columns)):
        raise ValueError(f"CSV incompleto. Debe incluir columnas: {sorted(req_cols)}")
    items = []
    for _, r in df.iterrows():
        items.append({
            "lat": float(r["lat"]),
            "lon": float(r["lon"]),
            "lat_round": float(r["lat_round"]),
            "lon_round": float(r["lon_round"]),
            "precision": int(r["precision"]),
            "plus_code": (str(r.get("plus_code")) if not pd.isna(r.get("plus_code")) else None),
            "nombre": (str(r.get("nombre")) if not pd.isna(r.get("nombre")) else None),
            "fuente": (str(r.get("fuente")) if not pd.isna(r.get("fuente")) else None),
            "confianza": (int(r.get("confianza")) if pd.notna(r.get("confianza")) else None),
        })
    repo_upsert_many(con, items)

# ===========================
# GEOCODING HTTP (sin geopy)
# ===========================
def _reverse_nominatim(lat, lon, lang="es") -> Tuple[str,str]:
    url = f"{NOMINATIM_URL}/reverse"
    params = {
        "format": "jsonv2",
        "lat": f"{float(lat):.6f}",
        "lon": f"{float(lon):.6f}",
        "accept-language": lang,
        "zoom": 18,
        "addressdetails": 1,
        "email": CONTACT_EMAIL
    }
    headers = {"User-Agent": USER_AGENT}
    r = requests.get(url, params=params, headers=headers, timeout=GEOCODER_TIMEOUT)
    r.raise_for_status()
    data = r.json()
    txt = data.get("display_name") if isinstance(data, dict) else ""
    return (txt or "", "nominatim")

def _reverse_locationiq(lat, lon, lang="es") -> Tuple[str,str]:
    if not LOCATIONIQ_KEY: return ("", "locationiq")
    url = "https://us1.locationiq.com/v1/reverse"
    params = {
        "key": LOCATIONIQ_KEY,
        "lat": f"{float(lat):.6f}",
        "lon": f"{float(lon):.6f}",
        "format": "json",
        "zoom": 18,
        "normalizeaddress": 1,
        "accept-language": lang
    }
    r = requests.get(url, params=params, timeout=GEOCODER_TIMEOUT)
    r.raise_for_status()
    js = r.json()
    txt = (js.get("display_name") or "") if isinstance(js, dict) else ""
    return (txt, "locationiq")

def _reverse_opencage(lat, lon, lang="es") -> Tuple[str,str]:
    if not OPENCAGE_KEY: return ("", "opencage")
    url = "https://api.opencagedata.com/geocode/v1/json"
    params = {
        "q": f"{float(lat):.6f},{float(lon):.6f}",
        "key": OPENCAGE_KEY,
        "language": lang,
        "no_annotations": 1,
        "limit": 1
    }
    r = requests.get(url, params=params, timeout=GEOCODER_TIMEOUT)
    r.raise_for_status()
    js = r.json()
    txt = js["results"][0].get("formatted", "") if (isinstance(js, dict) and js.get("results")) else ""
    return (txt, "opencage")

def _reverse_bigdatacloud(lat, lon, lang="es") -> Tuple[str,str]:
    url = "https://api.bigdatacloud.net/data/reverse-geocode-client"
    params = {
        "latitude": float(lat),
        "longitude": float(lon),
        "localityLanguage": "es" if str(lang).lower().startswith("es") else "en"
    }
    r = requests.get(url, params=params, timeout=GEOCODER_TIMEOUT)
    r.raise_for_status()
    js = r.json()
    if not isinstance(js, dict):
        return ("", "bigdatacloud")
    parts = [ js.get("locality") or js.get("city"), js.get("principalSubdivision"), js.get("countryName") ]
    out = ", ".join([p for p in parts if p])
    return (out or "", "bigdatacloud")

@lru_cache(maxsize=10000)
def reverse_best_cached(lat_round, lon_round, lang="es") -> Tuple[str,str]:
    for L in [lang, "es", "es-mx", "en"]:
        for fn in (_reverse_nominatim, _reverse_locationiq, _reverse_opencage, _reverse_bigdatacloud):
            try:
                txt, src = fn(lat_round, lon_round, L)
                if txt:
                    return (txt, src)
            except Exception:
                continue
    return ("", "")

def reverse_address_with_source(lat, lon, lang="es", precision=COORD_PRECISION_CACHE) -> Tuple[str,str]:
    lt = round(float(lat), precision); ln = round(float(lon), precision)
    return reverse_best_cached(lt, ln, lang)

# ===========================
# ESTADÍSTICOS (dtype-safe)
# ===========================
def build_statistics_blocks(df: pd.DataFrame) -> List[Tuple[str, pd.DataFrame]]:
    blocks: List[Tuple[str, pd.DataFrame]] = []
    if df is None or df.empty:
        return blocks

    df = df.copy()
    if "Es_Duplicado" in df.columns:
        df = df[~df["Es_Duplicado"].fillna(False)].copy()

    df["Tipo"] = df.get("Tipo").astype("string")
    df["Datetime"] = pd.to_datetime(df.get("Datetime"), errors="coerce")

    numA = df.get("Número A")
    numB = df.get("Número B")
    dur  = pd.to_numeric(df.get("Duración (seg)"), errors="coerce") if "Duración (seg)" in df.columns else pd.Series(dtype=float)

    def _clean_number(x: Any) -> str:
        if x is None or (isinstance(x, float) and pd.isna(x)): return ""
        return re.sub(r"\D+", "", str(x))

    tel = ""
    if "Teléfono" in df.columns:
        tel_mode = df["Teléfono"].map(_clean_number)
        tel = tel_mode.mode().iloc[0] if not tel_mode.dropna().empty else ""

    A_clean = numA.map(_clean_number) if numA is not None else pd.Series([""]*len(df))
    B_clean = numB.map(_clean_number) if numB is not None else pd.Series([""]*len(df))
    is_voice = df["Tipo"].astype(str).str.contains("VOZ", case=False, na=False)

    dt_valid = df["Datetime"].dropna().sort_values()
    rango_ini = dt_valid.min() if not dt_valid.empty else pd.NaT
    rango_fin = dt_valid.max() if not dt_valid.empty else pd.NaT
    dias_act  = dt_valid.dt.date.nunique() if not dt_valid.empty else 0
    max_gap_h = (dt_valid.diff().max().total_seconds()/3600) if (len(dt_valid) >= 2) else 0

    counts_by_tipo = df["Tipo"].value_counts(dropna=False).to_dict()
    voz_dur = dur[is_voice] if not dur.empty else pd.Series(dtype=float)

    resumen_rows = [
        ["Teléfono investigado", tel],
        ["Rango temporal - inicio", str(rango_ini) if pd.notna(rango_ini) else ""],
        ["Rango temporal - fin", str(rango_fin) if pd.notna(rango_fin) else ""],
        ["Días activos", dias_act],
        ["Máximo periodo de inactividad (h)", round(max_gap_h, 2)],
        ["Eventos totales", len(df)],
    ]
    for k, v in counts_by_tipo.items():
        resumen_rows.append([f"Total {k}", int(v)])
    if not voz_dur.empty:
        resumen_rows += [
            ["Duración VOZ total (s)", int(voz_dur.sum())],
            ["Duración VOZ media (s)", round(float(voz_dur.mean()), 2)],
            ["Duración VOZ mediana (s)", round(float(voz_dur.median()), 2)],
            ["Duración VOZ p95 (s)", round(float(voz_dur.quantile(0.95)), 2)],
        ]
    blocks.append(("1) Resumen global", pd.DataFrame(resumen_rows, columns=["Métrica","Valor"])))

    out_mask = (A_clean == tel) & is_voice
    in_mask  = (B_clean == tel) & is_voice

    dir_rows = []
    for nombre, m in [("SALIENTE", out_mask), ("ENTRANTE", in_mask)]:
        n = int(m.sum())
        dsum = int(dur[m].sum()) if not dur.empty else 0
        dmean = round(float(dur[m].mean()), 2) if n > 0 else 0.0
        dir_rows.append([nombre, n, dsum, dmean])
    blocks.append(("2) Dirección del tráfico (VOZ)", pd.DataFrame(dir_rows, columns=["Dirección","# Llamadas","Duración total (s)","Duración media (s)"])))

    if numB is not None:
        sal = pd.DataFrame(df[out_mask])
        if not sal.empty:
            sal = sal.assign(Contraparte=B_clean[out_mask].values)
            sal["Duración (seg)"] = pd.to_numeric(sal["Duración (seg)"], errors="coerce")
            top_sal = (sal.groupby("Contraparte", dropna=True)
                       .agg(Llamadas=("Datetime","count"),
                            Duración_total_s=("Duración (seg)","sum"),
                            Primera=("Datetime","min"),
                            Última=("Datetime","max"))
                       .reset_index().rename(columns={"Contraparte":"Número"}))
            top_sal = top_sal.sort_values(["Llamadas","Duración_total_s"], ascending=[False,False]).head(10)
            blocks.append(("3) Top 10 contactos SALIENTES", top_sal))
        else:
            blocks.append(("3) Top 10 contactos SALIENTES", pd.DataFrame(columns=["Número","Llamadas","Duración_total_s","Primera","Última"])))

    if numA is not None:
        ent = pd.DataFrame(df[in_mask])
        if not ent.empty:
            ent = ent.assign(Contraparte=A_clean[in_mask].values)
            ent["Duración (seg)"] = pd.to_numeric(ent["Duración (seg)"], errors="coerce")
            top_ent = (ent.groupby("Contraparte", dropna=True)
                       .agg(Llamadas=("Datetime","count"),
                            Duración_total_s=("Duración (seg)","sum"),
                            Primera=("Datetime","min"),
                            Última=("Datetime","max"))
                       ).reset_index().rename(columns={"Contraparte":"Número"})
            top_ent = top_ent.sort_values(["Llamadas","Duración_total_s"], ascending=[False,False]).head(10)
            blocks.append(("4) Top 10 contactos ENTRANTES", top_ent))
        else:
            blocks.append(("4) Top 10 contactos ENTRANTES", pd.DataFrame(columns=["Número","Llamadas","Duración_total_s","Primera","Última"])))

    if "PLUS_CODE_NOMBRE" in df.columns and df["PLUS_CODE_NOMBRE"].notna().any():
        antena_series = df["PLUS_CODE_NOMBRE"].astype(str).fillna("")
    else:
        lat = pd.to_numeric(df.get("Latitud"), errors="coerce") if "Latitud" in df.columns else pd.Series(index=df.index, dtype="float64")
        lon = pd.to_numeric(df.get("Longitud"), errors="coerce") if "Longitud" in df.columns else pd.Series(index=df.index, dtype="float64")
        antena_series = pd.Series("", index=df.index, dtype="object")
        mask_ok = lat.notna() & lon.notna()
        if mask_ok.any():
            antena_series.loc[mask_ok] = lat[mask_ok].round(6).astype(str) + "," + lon[mask_ok].round(6).astype(str)

    tipos_orden = ["DATOS", "VOZ ENTRANTE", "VOZ SALIENTE", "MENSAJES 2 VÍAS", "TRANSFER"]
    letras = ["A","B","C","D","E"]; idx = 0
    tipo_norm = df["Tipo"].astype(str).str.upper()

    for tnombre in tipos_orden:
        mask_t = tipo_norm.eq(tnombre)
        if not mask_t.any():
            continue
        tmp = pd.DataFrame({"Antena": antena_series[mask_t], "Datetime": df["Datetime"][mask_t]})
        tmp = tmp[tmp["Antena"].astype(str).str.len() > 0]
        if tmp.empty:
            continue
        top_ant_t = (tmp.groupby("Antena")
                     .agg(Eventos=("Antena","count"),
                          Primera=("Datetime","min"),
                          Última=("Datetime","max"))
                     ).reset_index().sort_values("Eventos", ascending=False).head(5)
        blocks.append((f"5{letras[idx]}) Antenas TOP — {tnombre}", top_ant_t)); idx += 1

    if "IMEI" in df.columns:
        imei_df = df.copy()
        imei_df["IMEI"] = imei_df["IMEI"].astype(str)
        imei_df = imei_df[imei_df["IMEI"].str.len() > 0]
        if not imei_df.empty:
            imei_tb = (imei_df.groupby("IMEI")
                       .agg(Eventos=("IMEI","count"),
                            Primera=("Datetime","min"),
                            Última=("Datetime","max"))
                       ).reset_index().sort_values("Eventos", ascending=False)
        else:
            imei_tb = pd.DataFrame(columns=["IMEI","Eventos","Primera","Última"])
    else:
        imei_tb = pd.DataFrame(columns=["IMEI","Eventos","Primera","Última"])
    blocks.append(("6) IMEI (uso por periodo)", imei_tb))

    return blocks

# ===========================
# UI CARGA
# ===========================
uploaded_files = st.file_uploader(
    "📂 Sube 1 o varias sábanas AT&T (CSV, XLS o XLSX)",
    type=["csv","xls","xlsx","txt"],
    accept_multiple_files=True
)

col1, col2 = st.columns(2)
go    = col1.button("🚀 Compilar y generar Excel (AT&T→Limpieza)", type="primary")
clear = col2.button("🗑️ Limpiar sesión")

if clear:
    try: st.rerun()
    except Exception: st.experimental_rerun()

# ===========================
# MAIN
# ===========================
if repo_export_btn:
    try:
        con = repo_connect(DB_PATH)
        df_repo = repo_export_df(con)
        con.close()
        csv_bytes = df_repo.to_csv(index=False).encode("utf-8")
        st.download_button("Descargar antenas_repo.csv", data=csv_bytes, file_name="antenas_repo.csv", mime="text/csv")
        st.success(f"Exportadas {len(df_repo):,} filas del repositorio.")
    except Exception as e:
        st.error("No se pudo exportar el repositorio.")
        st.exception(e)

if repo_import_file is not None:
    try:
        con = repo_connect(DB_PATH)
        df_imp = pd.read_csv(repo_import_file)
        repo_import_df(con, df_imp)
        con.close()
        st.success(f"Importadas/actualizadas {len(df_imp):,} filas al repositorio.")
    except Exception as e:
        st.error("No se pudo importar el CSV al repositorio.")
        st.exception(e)

if go:
    if not uploaded_files:
        st.warning("Primero sube al menos un archivo.")
        st.stop()

    progress = st.progress(0, text="Iniciando…")
    progress_section(progress, 4, "📥 Cargando archivos…")

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            paths: List[str] = []
            for f in uploaded_files:
                suffix = ("." + f.name.split(".")[-1].lower()) if "." in f.name else ""
                p = tempfile.NamedTemporaryFile(delete=False, dir=tmpdir, suffix=suffix).name
                with open(p, "wb") as w: w.write(f.getvalue())
                paths.append(p)

            # Leer y unir
            progress_section(progress, 12, "📄 Leyendo sábanas (header-sniff)…")
            frames, logs = [], []
            for p in paths:
                df = read_any_with_sniff(p)
                logs.append({"Archivo": os.path.basename(p), "Filas leídas": len(df), "Encabezados detectados": ", ".join(map(str, df.columns))})
                frames.append(df)
            raw_all = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

            # Transformación AT&T → 20 columnas
            progress_section(progress, 22, "🧩 Normalizando a 'Datos_Limpios (20 cols)'…")
            tel_fijo = telefono_fijo.strip() or None
            limpio = transform_att_to_limpio(raw_all, telefono_fijo=tel_fijo)

            # Resguardo originales geo + conversión DMS/decimal y corregir longitud negativa
            progress_section(progress, 30, "🧭 Convirtiendo coordenadas…")
            for col in ["Latitud","Longitud","Azimuth"]:
                if col in limpio.columns and f"{col}_raw" not in limpio.columns:
                    limpio[f"{col}_raw"] = limpio[col].astype(str)
            if "Latitud" in limpio.columns:
                limpio["Latitud"]  = limpio["Latitud"].apply(dms_to_decimal)
            if "Longitud" in limpio.columns:
                limpio["Longitud"] = limpio["Longitud"].apply(dms_to_decimal)
                limpio["Longitud"] = limpio["Longitud"].apply(lambda x: -abs(x) if pd.notna(x) else x)

            # PLUS_CODE offline si hay coords
            progress_section(progress, 40, "➕ Generando PLUS_CODE…")
            mask_coords = limpio["Latitud"].notna() & limpio["Longitud"].notna()
            if HAS_OLC and mask_coords.any():
                limpio.loc[mask_coords,"PLUS_CODE"] = limpio.loc[mask_coords].apply(
                    lambda x: plus_code(x["Latitud"], x["Longitud"]), axis=1
                )

            # ===== Resolver nombres con REPOSITORIO, luego geocodar lo faltante y actualizar repo =====
            progress_section(progress, 62, "📚 Resolviendo nombres de antenas…")
            if mask_coords.any():
                coords_unique = limpio.loc[mask_coords, ["Latitud","Longitud"]].drop_duplicates().reset_index(drop=True)
                coords_unique["lat_round"] = coords_unique["Latitud"].round(PRECISION)
                coords_unique["lon_round"] = coords_unique["Longitud"].round(PRECISION)

                resolved_map: Dict[Tuple[float,float], str] = {}
                plus_local: Dict[Tuple[float,float], str] = {}
                to_resolve: List[Tuple[float,float,float,float]] = []  # (lat,lon,lat_r,lon_r)

                if use_repo:
                    try:
                        con = repo_connect(DB_PATH)
                        repomap = repo_fetch_map(con, precision=PRECISION)
                        con.close()
                        for _, r in coords_unique.iterrows():
                            key = (float(r["lat_round"]), float(r["lon_round"]))
                            if key in repomap and repomap[key].get("nombre"):
                                resolved_map[key] = repomap[key]["nombre"]
                                if repomap[key].get("plus_code"): plus_local[key] = repomap[key]["plus_code"]
                            else:
                                to_resolve.append((float(r["Latitud"]), float(r["Longitud"]), key[0], key[1]))
                    except Exception as e:
                        st.warning(f"No se pudo abrir el repositorio: {e}")
                        for _, r in coords_unique.iterrows():
                            key = (float(r["lat_round"]), float(r["lon_round"]))
                            to_resolve.append((float(r["Latitud"]), float(r["Longitud"]), key[0], key[1]))
                else:
                    for _, r in coords_unique.iterrows():
                        key = (float(r["lat_round"]), float(r["lon_round"]))
                        to_resolve.append((float(r["Latitud"]), float(r["Longitud"]), key[0], key[1]))

                # Geocodar faltantes (con límite)
                geocoded_items: List[Dict[str, Any]] = []
                for idx, (lt, ln, ltr, lnr) in enumerate(to_resolve[:MAX_UNIQUE_GEOCODES], 1):
                    name, fuente = reverse_address_with_source(lt, ln, lang="es", precision=COORD_PRECISION_CACHE)
                    if not name:
                        name, fuente = reverse_address_with_source(lt, ln, lang="es-mx", precision=COORD_PRECISION_CACHE)
                    if not name:
                        name, fuente = reverse_address_with_source(lt, ln, lang="en", precision=COORD_PRECISION_CACHE)
                    if not name:
                        name, fuente = ("SIN_DIRECCIÓN", "")
                    key = (ltr, lnr)
                    resolved_map[key] = name
                    pcode = plus_code(lt, ln) if HAS_OLC else None
                    if pcode: plus_local[key] = pcode

                    if update_repo:
                        conf = {"nominatim":70, "locationiq":80, "opencage":85, "bigdatacloud":60}.get(fuente, None)
                        geocoded_items.append({
                            "lat": lt,
                            "lon": ln,
                            "lat_round": ltr,
                            "lon_round": lnr,
                            "precision": PRECISION,
                            "plus_code": pcode,
                            "nombre": name if name and name != "SIN_DIRECCIÓN" else None,
                            "fuente": fuente or None,
                            "confianza": conf,
                        })
                    pct = 62 + int(idx / max(len(to_resolve), 1) * (86 - 62))
                    progress_section(progress, pct, f"🌍 {idx}/{len(to_resolve)} coordenadas resueltas")

                if update_repo and geocoded_items:
                    try:
                        con = repo_connect(DB_PATH)
                        repo_upsert_many(con, geocoded_items)
                        con.close()
                        st.success(f"Repositorio actualizado con {len(geocoded_items)} nuevas coordenadas.")
                    except Exception as e:
                        st.warning(f"No fue posible actualizar el repositorio: {e}")

                # Volcar a DataFrame final
                def pick_name(lat, lon, curr):
                    if pd.notna(curr) and str(curr).strip():
                        return curr
                    ltr = round(float(lat), PRECISION); lnr = round(float(lon), PRECISION)
                    return resolved_map.get((ltr, lnr), "SIN_DIRECCIÓN")
                limpio.loc[mask_coords, "PLUS_CODE_NOMBRE"] = limpio.loc[mask_coords].apply(
                    lambda x: pick_name(x["Latitud"], x["Longitud"], x["PLUS_CODE_NOMBRE"]), axis=1
                )
                # Si faltaba PLUS_CODE y lo obtuvimos localmente
                if HAS_OLC:
                    def pick_plus(lat, lon, curr):
                        if pd.notna(curr) and str(curr).strip():
                            return curr
                        ltr = round(float(lat), PRECISION); lnr = round(float(lon), PRECISION)
                        return plus_local.get((ltr, lnr), None)
                    limpio.loc[mask_coords, "PLUS_CODE"] = limpio.loc[mask_coords].apply(
                        lambda x: pick_plus(x["Latitud"], x["Longitud"], x["PLUS_CODE"]), axis=1
                    )

            # Normalizar Tipo (estética)
            progress_section(progress, 90, "📚 Normalizando Tipo…")
            rep = {"voz entrante":"VOZ ENTRANTE","entrante":"VOZ ENTRANTE",
                   "voz saliente":"VOZ SALIENTE","saliente":"VOZ SALIENTE",
                   "datos":"DATOS","transfer":"TRANSFER",
                   "mensajes 2 vías":"MENSAJES 2 VÍAS","2 vías":"MENSAJES 2 VÍAS"}
            limpio["Tipo"] = limpio["Tipo"].astype(str).str.strip().str.lower().replace(rep).str.upper()

            # Marcar duplicados generales por (A,B,Datetime) para info
            subset_general = [c for c in ["Número A","Número B","Datetime"] if c in limpio.columns]
            if subset_general:
                dup_mask_all = limpio.duplicated(subset=subset_general, keep=False)
                limpio["Es_Duplicado"] = dup_mask_all
                key_col = subset_general[0]
                try:
                    limpio["Cuenta_GrupoDup"] = limpio.groupby(subset_general, dropna=False)[key_col].transform("size")
                except TypeError:
                    limpio["Cuenta_GrupoDup"] = limpio.groupby(subset_general)[key_col].transform("size")
            else:
                limpio["Es_Duplicado"] = False; limpio["Cuenta_GrupoDup"] = 1

            # Duplicados DATOS por minuto (opcional)
            progress_section(progress, 92, "🧽 Procesando duplicados (DATOS)…")
            duplicados_df = pd.DataFrame(); eliminados = 0
            if remove_dups:
                limpio, duplicados_df, eliminados = dedupe_datos_by_minute(limpio)

            # Orden final
            if "Datetime" in limpio.columns:
                limpio = limpio.sort_values(["Datetime","Número A","Número B"], ascending=[True,True,True], na_position="last").reset_index(drop=True)
            if not duplicados_df.empty and "Datetime" in duplicados_df.columns:
                duplicados_df = duplicados_df.sort_values(["Datetime","Número A","Número B"], ascending=[True,True,True], na_position="last").reset_index(drop=True)

            # LOG
            progress_section(progress, 94, "📜 Preparando LOG…")
            coords_validas    = int((limpio.get("Latitud", pd.Series(dtype=float)).notna() & limpio.get("Longitud", pd.Series(dtype=float)).notna()).sum())
            plus_generados    = int(limpio.get("PLUS_CODE", pd.Series(dtype=object)).notna().sum())
            nombres_generados = int(limpio.get("PLUS_CODE_NOMBRE", pd.Series(dtype=object)).notna().sum())
            log_df = pd.DataFrame({
                "Archivos procesados":[len(uploaded_files)],
                "Filas totales":[len(limpio)],
                "Duplicados eliminados (solo DATOS) ":[eliminados],
                "Coordenadas válidas detectadas":[coords_validas],
                "PLUS_CODE generados":[plus_generados],
                "PLUS_CODE_NOMBRE generados":[nombres_generados],
            })
            log_df2 = pd.DataFrame(logs)

            # ESTADÍSTICAS
            progress_section(progress, 96, "📊 Calculando estadísticas…")
            stat_blocks = build_statistics_blocks(limpio)

            # EXPORTAR EXCEL
            progress_section(progress, 98, "📦 Exportando a Excel…")
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                limpio.to_excel(writer, sheet_name="Datos_Limpios", index=False)
                log_df.to_excel(writer, sheet_name="LOG_Limpieza", index=False, startrow=0)
                log_df2.to_excel(writer, sheet_name="LOG_Limpieza", index=False, startrow=len(log_df)+2)
                if not duplicados_df.empty:
                    duplicados_df.to_excel(writer, sheet_name="Duplicados", index=False)
                # Hoja ESTADISTICAS
                start = 0; ws_title_positions: List[Tuple[int, str]] = []
                for title, bdf in stat_blocks:
                    bdf.to_excel(writer, sheet_name="ESTADISTICAS", index=False, startrow=start+1, startcol=0)
                    ws_title_positions.append((start+1, title))
                    start = start + 1 + 1 + len(bdf) + 1

            # Post format
            output.seek(0); wb = load_workbook(output)

            ws = wb["Datos_Limpios"]
            if "IMEI" in limpio.columns:
                imei_col = limpio.columns.get_loc("IMEI") + 1
                for row in ws.iter_rows(min_row=2, min_col=imei_col, max_col=imei_col):
                    for cell in row:
                        cell.number_format = "0"
            for col_idx, column in enumerate(ws.columns, start=1):
                max_len = 0; col_letter = get_column_letter(col_idx)
                for cell in column:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
            tab = Table(displayName="TablaLimpia", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(tab)

            if "LOG_Limpieza" in wb.sheetnames:
                ws_log = wb["LOG_Limpieza"]
                tab_log = Table(displayName="TablaLog", ref=f"A1:{get_column_letter(ws_log.max_column)}{ws_log.max_row}")
                tab_log.tableStyleInfo = TableStyleInfo(name="TableStyleLight11", showRowStripes=True)
                ws_log.add_table(tab_log)

            if "Duplicados" in wb.sheetnames:
                ws_dup = wb["Duplicados"]
                tab_dup = Table(displayName="TablaDuplicados", ref=f"A1:{get_column_letter(ws_dup.max_column)}{ws_dup.max_row}")
                tab_dup.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
                ws_dup.add_table(tab_dup)

            if "ESTADISTICAS" in wb.sheetnames:
                ws_est = wb["ESTADISTICAS"]
                for row_idx, title in ws_title_positions:
                    ws_est.cell(row=row_idx, column=1).value = title
                    ws_est.cell(row=row_idx, column=1).font = Font(bold=True)
                for col in ws_est.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    ws_est.column_dimensions[col_letter].width = min(max_len + 2, 60)

            final_output = io.BytesIO()
            wb.save(final_output); wb.close()
            final_output.seek(0)

            progress_section(progress, 100, "✅ Listo")

            st.success(f"✅ Hecho: {len(limpio):,} filas en 'Datos_Limpios' (20 columnas)")
            if show_preview:
                st.subheader("Preview — Datos_Limpios (20 columnas)")
                st.dataframe(limpio.head(500), use_container_width=True)

            st.subheader("📜 LOG")
            st.dataframe(log_df, use_container_width=True)
            st.dataframe(log_df2, use_container_width=True)

            st.subheader("📊 ESTADISTICAS")
            for title, bdf in stat_blocks:
                st.markdown(f"**{title}**")
                st.dataframe(bdf, use_container_width=True)

            st.download_button(
                "⬇️ Descargar archivo limpio",
                final_output.getvalue(),
                "archivo_limpio_geolocalizado.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"⚠️ Error durante la compilación: {e}")
        st.exception(e)
